import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config.database import get_db_connection

class CoretaxConverter:
    """Engine Converter dari Database SQLite ke Kertas Kerja Excel 6 Sub-Tab Lampiran L-1"""

    def __init__(self, wp_id):
        self.wp_id = wp_id

    def export_to_excel(self, output_filepath):
        """Membuat file Excel Kertas Kerja dengan 6 worksheet konsolidasi."""
        wb = openpyxl.Workbook()
        # Hapus sheet default 'Sheet'
        wb.remove(wb.active)

        # Style Definitions
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        font_title = Font(name="Calibri", size=14, bold=True)
        font_edited = Font(name="Calibri", size=11, color="9C0006", bold=True)
        fill_edited = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        conn = get_db_connection()
        cursor = conn.cursor()

        # Ambil data Master WP
        cursor.execute("SELECT npwp, nama_wp, tahun_pajak FROM master_wp WHERE id = ?", (self.wp_id,))
        wp_info = cursor.fetchone()
        if not wp_info:
            conn.close()
            raise ValueError("Data Wajib Pajak tidak ditemukan.")

        # Definisi Masing-Masing Sub-Tab Lampiran L-1
        categories = [
            {
                'tab_name': 'L1-Kas & Setara Kas',
                'table': 'harta_l1_kas',
                'headers': ['No', 'Kode Harta', 'Nama Harta', 'Tahun Perolehan', 'Harga Perolehan (IDR)', 'Keterangan'],
                'cols': ['kode_harta', 'nama_harta', 'tahun_perolehan', 'harga_perolehan', 'keterangan']
            },
            {
                'tab_name': 'L1-Piutang',
                'table': 'harta_l1_piutang',
                'headers': ['No', 'Kode Harta', 'Nama Peminjam', 'NPWP Peminjam', 'Tahun Perolehan', 'Harga Perolehan (IDR)', 'Keterangan'],
                'cols': ['kode_harta', 'nama_peminjam', 'npwp_peminjam', 'tahun_perolehan', 'harga_perolehan', 'keterangan']
            },
            {
                'tab_name': 'L1-Investasi',
                'table': 'harta_l1_investasi',
                'headers': ['No', 'Kode Harta', 'Nama Harta', 'Penerbit/Penyedia', 'Tahun Perolehan', 'Harga Perolehan (IDR)', 'Keterangan'],
                'cols': ['kode_harta', 'nama_harta', 'penerbit_saham', 'tahun_perolehan', 'harga_perolehan', 'keterangan']
            },
            {
                'tab_name': 'L1-Harta Bergerak',
                'table': 'harta_l1_bergerak',
                'headers': ['No', 'Kode Harta', 'Nama Harta', 'Merek / Tipe', 'Tahun Perolehan', 'Harga Perolehan (IDR)', 'Keterangan'],
                'cols': ['kode_harta', 'nama_harta', 'merek_type', 'tahun_perolehan', 'harga_perolehan', 'keterangan']
            },
            {
                'tab_name': 'L1-Harta Tidak Bergerak',
                'table': 'harta_l1_htb',
                'headers': ['No', 'Kode Harta', 'Jenis Harta', 'Lokasi / Alamat', 'Tahun Perolehan', 'Harga Perolehan (IDR)', 'Keterangan'],
                'cols': ['kode_harta', 'jenis_harta', 'lokasi_alamat', 'tahun_perolehan', 'harga_perolehan', 'keterangan']
            },
            {
                'tab_name': 'L1-Harta Lainnya',
                'table': 'harta_l1_lainnya',
                'headers': ['No', 'Kode Harta', 'Nama Harta', 'Tahun Perolehan', 'Harga Perolehan (IDR)', 'Keterangan'],
                'cols': ['kode_harta', 'nama_harta', 'tahun_perolehan', 'harga_perolehan', 'keterangan']
            }
        ]

        for cat in categories:
            ws = wb.create_sheet(title=cat['tab_name'])
            ws.views.sheetView[0].showGridLines = True

            # Header Metadata WP
            ws.cell(row=1, column=1, value=f"KERTAS KERJA KONSOLIDASI LAMPIRAN L-1 - {cat['tab_name'].upper()}").font = font_title
            ws.cell(row=2, column=1, value=f"Nama WP: {wp_info['nama_wp']} | NPWP: {wp_info['npwp']} | Tahun Pajak: {wp_info['tahun_pajak']}").font = Font(bold=True)

            # Table Header
            start_row = 4
            for col_num, header_title in enumerate(cat['headers'], 1):
                cell = ws.cell(row=start_row, column=col_num, value=header_title)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            # Query Data dari Database
            select_cols = ", ".join(cat['cols']) + ", is_edited"
            cursor.execute(f"SELECT {select_cols} FROM {cat['table']} WHERE wp_id = ?", (self.wp_id,))
            rows = cursor.fetchall()

            row_idx = start_row + 1
            for i, row in enumerate(rows, 1):
                ws.cell(row=row_idx, column=1, value=i).alignment = Alignment(horizontal="center")
                
                for col_idx, col_name in enumerate(cat['cols'], 2):
                    val = row[col_name]
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    
                    # Highlight cell jika data pernah diedit (Audit Trail Visual)
                    if row['is_edited'] == 1 and col_name == 'harga_perolehan':
                        cell.font = font_edited
                        cell.fill = fill_edited

                    # Formatting angka untuk Harga Perolehan
                    if col_name == 'harga_perolehan' and val is not None:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right")

                    cell.border = thin_border

                row_idx += 1

            # Auto-fit Column Widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        conn.close()
        
        # Simpan File
        os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
        wb.save(output_filepath)
        return output_filepath